# app.py
from flask import (
    Flask, render_template, render_template_string, request, redirect, url_for, flash, session,
    send_file, send_from_directory, jsonify, Response
)
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from datetime import datetime, timedelta
from io import BytesIO
from sqlalchemy import text, func, Index
from werkzeug.middleware.proxy_fix import ProxyFix
from jinja2 import TemplateNotFound
import os
import time
import hashlib

# ========= APP / CONFIG =========
app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "coopex-secreto")

# Corrige scheme/host atrás do proxy para cookies seguros e redirects corretos
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_port=1)

# Sessão/Cookies
app.permanent_session_lifetime = timedelta(hours=10)
app.config.update(
    SESSION_COOKIE_SAMESITE='Lax',
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SECURE=True,   # Render usa HTTPS
    TEMPLATES_AUTO_RELOAD=True,   # útil no dev
    JSONIFY_PRETTYPRINT_REGULAR=False,
    JSON_SORT_KEYS=False,
)

def _build_db_uri() -> str:
    """
    Usa DATABASE_URL se existir; senão, mantém um fallback.
    Converte postgres:// -> postgresql+psycopg:// e aplica sslmode=require.
    """
    url = os.environ.get("DATABASE_URL")
    if not url:
        return (
            'postgresql+psycopg://'
            'banco_dados_9ooo_user:4eebYkKJwygTnOzrU1PAMFphnIli4iCH'
            '@dpg-d28sr2juibrs73du5n80-a.oregon-postgres.render.com/banco_dados_9ooo'
            '?sslmode=require'
        )
    if url.startswith("postgres://"):
        url = url.replace("postgres://", "postgresql+psycopg://", 1)
    elif url.startswith("postgresql://") and "+psycopg" not in url:
        url = url.replace("postgresql://", "postgresql+psycopg://", 1)
    if url.startswith("postgresql+psycopg://") and "sslmode=" not in url:
        url += ("&" if "?" in url else "?") + "sslmode=require"
    return url

# SQLAlchemy
app.config['SQLALCHEMY_DATABASE_URI'] = _build_db_uri()
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    'pool_pre_ping': True,
    'pool_size': 5,
    'max_overflow': 10,
    'pool_timeout': 30,
    'pool_recycle': 1800,
}

# Estáticos (cache padrão de 1 dia)
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 24 * 60 * 60

# Pastas
app.config['UPLOAD_FOLDER_COOPERADOS'] = 'static/uploads'
app.config['UPLOAD_FOLDER_LOGOS'] = 'static/logos'
app.config['STATICS_FOLDER'] = 'statics'
os.makedirs(app.config['UPLOAD_FOLDER_COOPERADOS'], exist_ok=True)
os.makedirs(app.config['UPLOAD_FOLDER_LOGOS'], exist_ok=True)
os.makedirs(app.config['STATICS_FOLDER'], exist_ok=True)

# Compressão Gzip (opcional)
try:
    from flask_compress import Compress
    Compress(app)
except Exception:
    pass

db = SQLAlchemy(app)

# ====== garantir criação de tabelas em runtime (sem apagar nada) ======
def ensure_schema():
    """Cria colunas no banco se ainda não existirem (sem Alembic)."""
    with app.app_context():
        try:
            cols = {r[0] for r in db.session.execute(text(
                "SELECT column_name FROM information_schema.columns WHERE table_name = 'cooperado'"
            )).fetchall()}
        except Exception:
            # Se não for Postgres (ex.: SQLite), simplesmente não aplicamos ALTERs
            cols = set()

        alter_stmts = []
        if 'foto_data' not in cols:
            alter_stmts.append("ADD COLUMN IF NOT EXISTS foto_data BYTEA")
        if 'foto_mimetype' not in cols:
            alter_stmts.append("ADD COLUMN IF NOT EXISTS foto_mimetype VARCHAR(50)")
        if 'foto_filename' not in cols:
            alter_stmts.append("ADD COLUMN IF NOT EXISTS foto_filename VARCHAR(120)")
        if 'credito_atualizado_em' not in cols:
            alter_stmts.append("ADD COLUMN IF NOT EXISTS credito_atualizado_em TIMESTAMP NULL")
        if 'senha_hash' not in cols:
            alter_stmts.append("ADD COLUMN IF NOT EXISTS senha_hash VARCHAR(128)")

        if alter_stmts and 'information_schema' in _build_db_uri():
            # proteção simples: só tenta no Postgres
            try:
                db.session.execute(text("ALTER TABLE cooperado " + ", ".join(alter_stmts)))
                db.session.commit()
            except Exception:
                db.session.rollback()

try:
    with app.app_context():
        db.create_all()
        ensure_schema()
except Exception:
    pass

# ========= MODELS =========
class Cooperado(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(120), nullable=False)
    username = db.Column(db.String(80), unique=True, nullable=False, index=True)
    credito = db.Column(db.Float, default=0)
    credito_atualizado_em = db.Column(db.DateTime, index=True)
    # foto (banco e legado no disco)
    foto = db.Column(db.String(120), nullable=True)
    foto_data = db.Column(db.LargeBinary, nullable=True)      # BYTEA
    foto_mimetype = db.Column(db.String(50), nullable=True)
    foto_filename = db.Column(db.String(120), nullable=True)
    # senha do cooperado
    senha_hash = db.Column(db.String(128), nullable=True)

    def set_senha(self, senha: str):
        self.senha_hash = generate_password_hash(senha)

    def checar_senha(self, senha: str) -> bool:
        return bool(self.senha_hash) and check_password_hash(self.senha_hash, senha)

class Estabelecimento(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(120), nullable=False)
    username = db.Column(db.String(80), unique=True, nullable=False, index=True)
    senha_hash = db.Column(db.String(128), nullable=False)
    logo = db.Column(db.String(120), nullable=True)
    def set_senha(self, senha):
        self.senha_hash = generate_password_hash(senha)
    def checar_senha(self, senha):
        return check_password_hash(self.senha_hash, senha)

class Admin(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(120), nullable=False)
    username = db.Column(db.String(80), unique=True, nullable=False, index=True)
    senha_hash = db.Column(db.String(128), nullable=False)
    def set_senha(self, senha):
        self.senha_hash = generate_password_hash(senha)
    def checar_senha(self, senha):
        return check_password_hash(self.senha_hash, senha)

class Lancamento(db.Model):
    __tablename__ = 'lancamento'
    id = db.Column(db.Integer, primary_key=True)
    data = db.Column(db.DateTime, nullable=False, default=datetime.utcnow, index=True)
    os_numero = db.Column(db.String(50), nullable=False)
    cooperado_id = db.Column(db.Integer, db.ForeignKey('cooperado.id'), nullable=False, index=True)
    estabelecimento_id = db.Column(db.Integer, db.ForeignKey('estabelecimento.id'), nullable=False, index=True)
    valor = db.Column(db.Float, nullable=False)
    descricao = db.Column(db.String(250))
    cooperado = db.relationship('Cooperado')
    estabelecimento = db.relationship('Estabelecimento')

Index('ix_lancamento_coop_estab_data', Lancamento.cooperado_id, Lancamento.estabelecimento_id, Lancamento.data.desc())

# ========= CONTEXT PROCESSOR (corrige uso de now()/callable no Jinja) =========
@app.context_processor
def inject_globals():
    return {
        "now": datetime.now,           # permite {{ now() }}
        "current_year": datetime.now().year,
        "callable": callable           # se o template usar {{ callable(now) }}
    }

# ========= HELPERS =========
def is_admin():
    return session.get('user_tipo') == 'admin'

def is_estabelecimento():
    return session.get('user_tipo') == 'estabelecimento'

def is_cooperado():
    return session.get('user_tipo') == 'cooperado'

def parse_date(s):
    if not s:
        return None
    try:
        if len(s.strip()) <= 10:
            return datetime.strptime(s, '%Y-%m-%d')
        return datetime.strptime(s, '%Y-%m-%d %H:%M')
    except Exception:
        return None

def _cache_headers(seconds=None, etag_base=None):
    """Gera cabeçalhos Cache-Control e ETag simples."""
    if seconds is None:
        seconds = int(app.config.get('SEND_FILE_MAX_AGE_DEFAULT', 3600))
    headers = {
        "Cache-Control": f"public, max-age={seconds}, immutable" if seconds >= 86400 else f"public, max-age={seconds}"
    }
    if etag_base:
        etag = hashlib.sha256(etag_base.encode('utf-8')).hexdigest()[:16]
        headers["ETag"] = etag
    return headers

def _response_with_cache(resp: Response, seconds=None, etag_base=None):
    headers = _cache_headers(seconds, etag_base)
    for k, v in headers.items():
        resp.headers[k] = v
    return resp

# ========= CACHE LEVE para /api/ultimo_lancamento =========
_LAST_LANC_CACHE = {"value": 0, "ts": 0.0}
_LAST_LANC_TTL = 2.0  # segundos

def _get_cached_last_lanc_id():
    now_ts = time.time()
    if now_ts - _LAST_LANC_CACHE["ts"] <= _LAST_LANC_TTL and _LAST_LANC_CACHE["ts"] > 0:
        return _LAST_LANC_CACHE["value"], True
    last_id = db.session.query(func.max(Lancamento.id)).scalar() or 0
    _LAST_LANC_CACHE["value"] = int(last_id)
    _LAST_LANC_CACHE["ts"] = now_ts
    return _LAST_LANC_CACHE["value"], False

def _invalidate_last_lanc_cache():
    _LAST_LANC_CACHE["ts"] = 0.0

def _update_last_lanc_cache_with_value(v: int):
    _LAST_LANC_CACHE["value"] = max(_LAST_LANC_CACHE["value"], int(v))
    _LAST_LANC_CACHE["ts"] = time.time()

# ========= ESTÁTICOS =========
@app.route('/statics/<path:filename>')
def statics_files(filename):
    resp = send_from_directory(app.config['STATICS_FOLDER'], filename)
    return _response_with_cache(resp, etag_base=f"statics/{filename}")

# ========= HEADERS GERAIS DE PERFORMANCE =========
@app.after_request
def add_perf_headers(resp: Response):
    resp.headers.setdefault("Connection", "keep-alive")
    resp.headers.setdefault("Server-Timing", "app;desc=\"Coopex-API\"")
    return resp

# ========= LOGIN/LOGOUT =========
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        tipo = request.form.get('tipo', '').strip()
        username = request.form.get('username', '').strip()
        senha = request.form.get('senha', '')
        session.permanent = True

        if tipo == 'admin':
            user = Admin.query.filter_by(username=username).first()
            if user and user.checar_senha(senha):
                session['user_id'] = user.id
                session['user_tipo'] = 'admin'
                return redirect(url_for('dashboard'))

        elif tipo == 'estabelecimento':
            est = Estabelecimento.query.filter_by(username=username).first()
            if est and est.checar_senha(senha):
                session['user_id'] = est.id
                session['user_tipo'] = 'estabelecimento'
                return redirect(url_for('painel_estabelecimento'))

        elif tipo == 'cooperado':
            coop = Cooperado.query.filter_by(username=username).first()
            if coop and coop.checar_senha(senha):
                session['user_id'] = coop.id
                session['user_tipo'] = 'cooperado'
                return redirect(url_for('painel_cooperado'))

        else:
            flash('Selecione o tipo de usuário.', 'danger')
            return render_template('login.html'), 400

        flash('Usuário ou senha inválidos', 'danger')
        return render_template('login.html'), 401

    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

# ========= DASHBOARD (ADMIN) =========
@app.route('/')
@app.route('/dashboard')
def dashboard():
    if not is_admin():
        return redirect(url_for('login'))
    admin = Admin.query.get(session['user_id'])
    cooperados = Cooperado.query.order_by(Cooperado.nome).all()
    estabelecimentos = Estabelecimento.query.order_by(Estabelecimento.nome).all()

    filtros = {
        'cooperado_id': request.args.get('cooperado_id'),
        'estabelecimento_id': request.args.get('estabelecimento_id'),
        'data_inicio': request.args.get('data_inicio'),
        'data_fim': request.args.get('data_fim')
    }
    coop_id_i = int(filtros['cooperado_id']) if filtros['cooperado_id'] else None
    est_id_i  = int(filtros['estabelecimento_id']) if filtros['estabelecimento_id'] else None
    di = parse_date(filtros['data_inicio'])
    df = parse_date(filtros['data_fim'])

    base_q = db.session.query(Lancamento)
    if coop_id_i is not None:
        base_q = base_q.filter(Lancamento.cooperado_id == coop_id_i)
    if est_id_i is not None:
        base_q = base_q.filter(Lancamento.estabelecimento_id == est_id_i)
    if di:
        base_q = base_q.filter(Lancamento.data >= di)
    if df:
        base_q = base_q.filter(Lancamento.data <= df)

    total_pedidos = base_q.count()
    sum_q = db.session.query(func.coalesce(func.sum(Lancamento.valor), 0.0))
    if coop_id_i is not None:
        sum_q = sum_q.filter(Lancamento.cooperado_id == coop_id_i)
    if est_id_i is not None:
        sum_q = sum_q.filter(Lancamento.estabelecimento_id == est_id_i)
    if di:
        sum_q = sum_q.filter(Lancamento.data >= di)
    if df:
        sum_q = sum_q.filter(Lancamento.data <= df)
    total_valor = (sum_q.scalar() or 0.0)

    sum_per_coop = db.session.query(
        Cooperado.id,
        Cooperado.nome,
        func.coalesce(func.sum(Lancamento.valor), 0.0).label('total')
    ).outerjoin(
        Lancamento, Cooperado.id == Lancamento.cooperado_id
    )

    if coop_id_i is not None:
        sum_per_coop = sum_per_coop.filter(Cooperado.id == coop_id_i)
    if est_id_i is not None:
        sum_per_coop = sum_per_coop.filter((Lancamento.estabelecimento_id == est_id_i) | (Lancamento.id.is_(None)))
    if di:
        sum_per_coop = sum_per_coop.filter((Lancamento.data >= di) | (Lancamento.id.is_(None)))
    if df:
        sum_per_coop = sum_per_coop.filter((Lancamento.data <= df) | (Lancamento.id.is_(None)))

    sum_per_coop = sum_per_coop.group_by(Cooperado.id, Cooperado.nome).order_by(Cooperado.nome).all()

    cooperado_nomes = [row.nome for row in sum_per_coop] or ["Nenhum cooperado"]
    cooperado_valores = [float(row.total) for row in sum_per_coop] or [0.0]

    try:
        ultimo_lancamento_id, _ = _get_cached_last_lanc_id()
    except Exception:
        ultimo_lancamento_id = 0

    return render_template('dashboard.html',
        admin=admin,
        cooperados=cooperados,
        estabelecimentos=estabelecimentos,
        total_pedidos=total_pedidos,
        total_valor=total_valor,
        total_cooperados=len(cooperados),
        total_estabelecimentos=len(estabelecimentos),
        cooperado_nomes=cooperado_nomes,
        cooperado_valores=cooperado_valores,
        lancamentos_contagem=cooperado_valores,
        filtros=filtros,
        ultimo_lancamento_id=ultimo_lancamento_id
    )

@app.route('/painel_admin')
def painel_admin():
    return redirect(url_for('dashboard'))

# ========= APIs para o Dashboard detectar novos lançamentos =========
@app.get('/api/ultimo_lancamento')
def api_ultimo_lancamento():
    last_id, cached = _get_cached_last_lanc_id()
    resp = jsonify({"last_id": int(last_id)})
    resp.headers['Cache-Control'] = 'public, max-age=2'
    if cached:
        resp.headers['X-Cache-Hit'] = '1'
    return resp

@app.get('/api/lancamento_info')
def api_lancamento_info():
    lanc_id = request.args.get('id', type=int)
    if not lanc_id:
        return jsonify({"error": "id requerido"}), 400
    l = Lancamento.query.get_or_404(lanc_id)
    nome = l.cooperado.nome if l.cooperado else ""
    return jsonify({
        "id": l.id,
        "cooperado": nome,
        "valor": float(l.valor),
        "os_numero": l.os_numero
    })

# ========= COOPERADOS CRUD =========
@app.route('/listar_cooperados')
def listar_cooperados():
    if not is_admin():
        return redirect(url_for('login'))
    admin = Admin.query.get(session['user_id'])
    cooperados = Cooperado.query.order_by(Cooperado.nome).all()
    return render_template('cooperados.html', admin=admin, cooperados=cooperados)

@app.route('/cooperados/novo', methods=['GET', 'POST'])
def novo_cooperado():
    if not is_admin():
        return redirect(url_for('login'))
    if request.method == 'POST':
        nome = request.form['nome']
        username = request.form['username'].strip()
        credito = float(request.form.get('credito', 0) or 0)

        # senha (obrigatória no cadastro)
        senha = (request.form.get('senha') or '').strip()
        senha2 = (request.form.get('senha2') or '').strip()
        if not senha or senha != senha2:
            flash('Defina a senha e confirme corretamente.', 'danger')
            return redirect(url_for('novo_cooperado'))

        # foto
        foto_file = request.files.get('foto')
        foto_filename = None
        foto_data = None
        foto_mimetype = None
        if foto_file and foto_file.filename:
            foto_filename = secure_filename(f"foto_{username}_{foto_file.filename}")
            foto_file.stream.seek(0)
            raw = foto_file.read()
            foto_data = raw
            foto_mimetype = foto_file.mimetype
            try:
                with open(os.path.join(app.config['UPLOAD_FOLDER_COOPERADOS'], foto_filename), 'wb') as f:
                    f.write(raw)
            except Exception:
                pass

        if Cooperado.query.filter_by(username=username).first():
            flash('Usuário já existe!', 'danger')
            return redirect(url_for('novo_cooperado'))

        cooperado = Cooperado(
            nome=nome, username=username, credito=credito,
            foto=foto_filename, foto_data=foto_data,
            foto_mimetype=foto_mimetype, foto_filename=foto_filename
        )
        cooperado.set_senha(senha)

        db.session.add(cooperado)
        db.session.commit()
        flash('Cooperado cadastrado!', 'success')
        return redirect(url_for('listar_cooperados'))
    return render_template('cooperado_form.html', editar=False, cooperado=None)

@app.route('/cooperados/editar/<int:id>', methods=['GET', 'POST'])
def editar_cooperado(id):
    if not is_admin():
        return redirect(url_for('login'))
    cooperado = Cooperado.query.get_or_404(id)
    if request.method == 'POST':
        cooperado.nome = request.form['nome']

        # atualizar username se mudou (checando unicidade)
        new_username = request.form.get('username', cooperado.username).strip()
        if new_username != cooperado.username:
            if Cooperado.query.filter(Cooperado.username == new_username, Cooperado.id != cooperado.id).first():
                flash('Este usuário já está em uso.', 'danger')
                return redirect(url_for('editar_cooperado', id=id))
            cooperado.username = new_username

        # crédito
        if 'credito' in request.form:
            try:
                novo_credito = float(request.form.get('credito', cooperado.credito) or cooperado.credito)
                if novo_credito != cooperado.credito:
                    cooperado.credito = novo_credito
                    cooperado.credito_atualizado_em = datetime.utcnow()
            except Exception:
                flash('Crédito inválido.', 'danger')
                return redirect(url_for('editar_cooperado', id=id))

        # foto
        foto_file = request.files.get('foto')
        if foto_file and foto_file.filename:
            foto_filename = secure_filename(f"foto_{cooperado.username}_{foto_file.filename}")
            foto_file.stream.seek(0)
            raw = foto_file.read()
            cooperado.foto = foto_filename
            cooperado.foto_filename = foto_filename
            cooperado.foto_data = raw
            cooperado.foto_mimetype = foto_file.mimetype
            try:
                with open(os.path.join(app.config['UPLOAD_FOLDER_COOPERADOS'], foto_filename), 'wb') as f:
                    f.write(raw)
            except Exception:
                pass

        # trocar senha se informado
        senha = (request.form.get('senha') or '').strip()
        senha2 = (request.form.get('senha2') or '').strip()
        if senha or senha2:
            if senha != senha2:
                flash('As senhas não conferem.', 'danger')
                return redirect(url_for('editar_cooperado', id=id))
            cooperado.set_senha(senha)

        db.session.commit()
        flash('Cooperado alterado!', 'success')
        return redirect(url_for('listar_cooperados'))
    return render_template('cooperado_form.html', editar=True, cooperado=cooperado)

@app.route('/cooperados/excluir/<int:id>')
def excluir_cooperado(id):
    if not is_admin():
        return redirect(url_for('login'))
    cooperado = Cooperado.query.get_or_404(id)
    db.session.delete(cooperado)
    db.session.commit()
    flash('Cooperado excluído!', 'success')
    return redirect(url_for('listar_cooperados'))

# Serve foto do cooperado (prioriza banco; fallback no disco) com cache correto
@app.route('/cooperados/foto/<int:id>')
def foto_cooperado(id):
    c = Cooperado.query.get_or_404(id)
    cache_sec = int(app.config.get('SEND_FILE_MAX_AGE_DEFAULT', 86400))

    if c.foto_data:
        bio = BytesIO(c.foto_data)
        resp = send_file(
            bio,
            mimetype=c.foto_mimetype or 'image/jpeg',
            download_name=c.foto_filename or f'cooperado_{id}.jpg'
        )
        return _response_with_cache(resp, cache_sec, etag_base=f"cooperado_db_{id}_{len(c.foto_data)}")

    if c.foto:
        path = os.path.join(app.config['UPLOAD_FOLDER_COOPERADOS'], c.foto)
        if os.path.exists(path):
            resp = send_file(path, mimetype='image/jpeg')
            try:
                size = os.path.getsize(path)
            except Exception:
                size = 0
            return _response_with_cache(resp, cache_sec, etag_base=f"cooperado_fs_{id}_{size}")

    resp = send_file(BytesIO(b''), mimetype='image/jpeg')
    return _response_with_cache(resp, cache_sec, etag_base=f"cooperado_empty_{id}")

# ========= AJUSTAR CRÉDITO =========
@app.route('/ajustar_credito', methods=['GET', 'POST'])
def ajustar_credito():
    if not is_admin():
        return redirect(url_for('login'))
    cooperado_id = request.args.get('id')
    if cooperado_id:
        return ajustar_credito_individual(int(cooperado_id))

    cooperados = Cooperado.query.order_by(Cooperado.nome).all()
    if request.method == 'POST':
        cooperado_id = request.form.get('cooperado_id')
        novo_credito = request.form.get('credito')
        if cooperado_id and novo_credito is not None:
            c = Cooperado.query.get(int(cooperado_id))
            if c:
                try:
                    c.credito = float(novo_credito)
                except Exception:
                    flash('Crédito inválido.', 'danger')
                    return redirect(url_for('ajustar_credito'))
                c.credito_atualizado_em = datetime.utcnow()
                db.session.commit()
                flash('Crédito ajustado!', 'success')
                return redirect(url_for('ajustar_credito'))
            else:
                flash('Cooperado não encontrado!', 'danger')
        else:
            flash('Selecione um cooperado e valor.', 'danger')
    return render_template('ajustar_credito.html', cooperados=cooperados)

@app.route('/ajustar_credito/<int:id>', methods=['GET', 'POST'])
def ajustar_credito_individual(id):
    if not is_admin():
        return redirect(url_for('login'))
    cooperado = Cooperado.query.get_or_404(id)
    if request.method == 'POST':
        novo_credito = request.form.get('credito')
        if novo_credito is not None:
            try:
                cooperado.credito = float(novo_credito)
            except Exception:
                flash('Crédito inválido.', 'danger')
                return redirect(url_for('ajustar_credito_individual', id=id))
            cooperado.credito_atualizado_em = datetime.utcnow()
            db.session.commit()
            flash('Crédito ajustado!', 'success')
            return redirect(url_for('listar_cooperados'))
    return render_template('ajustar_credito.html', cooperado=cooperado)

# ========= ESTABELECIMENTOS CRUD =========
@app.route('/listar_estabelecimentos')
def listar_estabelecimentos():
    if not is_admin():
        return redirect(url_for('login'))
    admin = Admin.query.get(session['user_id'])
    estabelecimentos = Estabelecimento.query.order_by(Estabelecimento.nome).all()
    return render_template('estabelecimentos.html', admin=admin, estabelecimentos=estabelecimentos)

@app.route('/novo_estabelecimento', methods=['GET', 'POST'])
def novo_estabelecimento():
    if not is_admin():
        return redirect(url_for('login'))
    if request.method == 'POST':
        nome = request.form['nome']
        username = request.form['username'].strip()
        senha = request.form['senha']
        logo_file = request.files.get('logo')
        filename = None
        if logo_file and logo_file.filename:
            filename = secure_filename(logo_file.filename)
            logo_file.save(os.path.join(app.config['UPLOAD_FOLDER_LOGOS'], filename))
        if Estabelecimento.query.filter_by(username=username).first():
            flash('Usuário já existe!', 'danger')
            return redirect(url_for('novo_estabelecimento'))
        est = Estabelecimento(nome=nome, username=username, logo=filename)
        est.set_senha(senha)
        db.session.add(est)
        db.session.commit()
        flash('Estabelecimento cadastrado!', 'success')
        return redirect(url_for('listar_estabelecimentos'))
    return render_template('estabelecimento_form.html', editar=False, estabelecimento=None)

@app.route('/editar_estabelecimento/<int:id>', methods=['GET', 'POST'])
def editar_estabelecimento(id):
    if not is_admin():
        return redirect(url_for('login'))
    est = Estabelecimento.query.get_or_404(id)
    if request.method == 'POST':
        est.nome = request.form['nome']
        if request.form['senha']:
            est.set_senha(request.form['senha'])
        logo_file = request.files.get('logo')
        if logo_file and logo_file.filename:
            filename = secure_filename(logo_file.filename)
            logo_file.save(os.path.join(app.config['UPLOAD_FOLDER_LOGOS'], filename))
            est.logo = filename
        db.session.commit()
        flash('Estabelecimento alterado!', 'success')
        return redirect(url_for('listar_estabelecimentos'))
    return render_template('estabelecimento_form.html', editar=True, estabelecimento=est)

@app.route('/excluir_estabelecimento/<int:id>')
def excluir_estabelecimento(id):
    if not is_admin():
        return redirect(url_for('login'))
    est = Estabelecimento.query.get_or_404(id)
    db.session.delete(est)
    db.session.commit()
    flash('Estabelecimento excluído!', 'success')
    return redirect(url_for('listar_estabelecimentos'))

# ========= LANÇAMENTOS (ADMIN) =========
@app.route('/lancamentos')
def listar_lancamentos():
    if not is_admin():
        return redirect(url_for('login'))
    admin = Admin.query.get(session['user_id'])
    cooperados = Cooperado.query.order_by(Cooperado.nome).all()
    estabelecimentos = Estabelecimento.query.order_by(Estabelecimento.nome).all()
    filtros = {
        'cooperado_id': request.args.get('cooperado_id'),
        'estabelecimento_id': request.args.get('estabelecimento_id'),
        'data_inicio': request.args.get('data_inicio'),
        'data_fim': request.args.get('data_fim')
    }
    coop_id_i = int(filtros['cooperado_id']) if filtros['cooperado_id'] else None
    est_id_i  = int(filtros['estabelecimento_id']) if filtros['estabelecimento_id'] else None
    di = parse_date(filtros['data_inicio'])
    df = parse_date(filtros['data_fim'])

    query = Lancamento.query
    if coop_id_i is not None:
        query = query.filter(Lancamento.cooperado_id == coop_id_i)
    if est_id_i is not None:
        query = query.filter(Lancamento.estabelecimento_id == est_id_i)
    if di:
        query = query.filter(Lancamento.data >= di)
    if df:
        query = query.filter(Lancamento.data <= df)

    lancamentos = query.order_by(Lancamento.data.desc()).all()
    return render_template('lancamentos.html',
                           admin=admin,
                           cooperados=cooperados,
                           estabelecimentos=estabelecimentos,
                           lancamentos=lancamentos,
                           filtros=filtros)

@app.route('/lancamentos/exportar')
def exportar_lancamentos():
    if not is_admin():
        return redirect(url_for('login'))

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return ("Para exportar, inclua 'openpyxl>=3.1.2' no requirements.txt e redeploy."), 500

    coop_id = request.args.get('cooperado_id')
    est_id = request.args.get('estabelecimento_id')
    di_s = request.args.get('data_inicio')
    df_s = request.args.get('data_fim')

    coop_id_i = int(coop_id) if coop_id else None
    est_id_i  = int(est_id) if est_id else None
    di = parse_date(di_s)
    df = parse_date(df_s)

    q = Lancamento.query
    if coop_id_i is not None:
        q = q.filter(Lancamento.cooperado_id == coop_id_i)
    if est_id_i is not None:
        q = q.filter(Lancamento.estabelecimento_id == est_id_i)
    if di:
        q = q.filter(Lancamento.data >= di)
    if df:
        q = q.filter(Lancamento.data <= df)
    q = q.order_by(Lancamento.data.desc())

    rows = q.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Lançamentos"

    header = ["Data", "Nº OS", "Cooperado", "Estabelecimento", "Valor (R$)", "Descrição"]
    ws.append(header)
    for col_idx, h in enumerate(header, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for l in rows:
        coop_nome = l.cooperado.nome if l.cooperado else ""
        est_nome = l.estabelecimento.nome if l.estabelecimento else ""
        data_fmt = l.data.strftime('%d/%m/%Y %H:%M')
        ws.append([data_fmt, l.os_numero, coop_nome, est_nome, float(l.valor), l.descricao or ""])

    widths = [20, 16, 32, 32, 16, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=5).number_format = u'"R$" #,##0.00'

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    resp = send_file(
        bio,
        as_attachment=True,
        download_name="lancamentos.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    return _response_with_cache(resp, 0, etag_base=f"xlsx_{len(rows)}")

# ====== EXCLUSÃO DE LANÇAMENTO (ADMIN, SEM LIMITE) ======
@app.post('/lancamentos/<int:id>/excluir')
def excluir_lancamento(id):
    if not is_admin():
        return redirect(url_for('login'))

    l = Lancamento.query.get_or_404(id)

    # Devolve o crédito ao cooperado
    cooperado = Cooperado.query.get(l.cooperado_id)
    if cooperado:
        cooperado.credito += float(l.valor or 0)

    try:
        db.session.delete(l)
        db.session.commit()
        _invalidate_last_lanc_cache()
        flash('Lançamento excluído e crédito devolvido ao cooperado.', 'success')
    except Exception:
        db.session.rollback()
        flash('Não foi possível excluir o lançamento.', 'danger')

    next_url = request.form.get('next') or url_for('listar_lancamentos')
    return redirect(next_url)

# ========= PAINEL ESTABELECIMENTO =========
@app.route('/painel_estabelecimento', methods=['GET', 'POST'])
def painel_estabelecimento():
    if not is_estabelecimento():
        return redirect(url_for('login'))
    est = Estabelecimento.query.get(session['user_id'])
    cooperados = Cooperado.query.order_by(Cooperado.nome).all()

    if request.method == 'POST':
        cooperado_id = request.form.get('cooperado_id')
        valor = request.form.get('valor')
        os_numero = request.form.get('os_numero')
        descricao = request.form.get('descricao')

        if cooperado_id and valor and os_numero:
            c = Cooperado.query.get(int(cooperado_id))
            if c:
                valor_f = float(valor)
                novo_credito = c.credito - valor_f
                if novo_credito < 0:
                    flash('Crédito insuficiente para este lançamento.', 'danger')
                else:
                    l = Lancamento(
                        data=datetime.utcnow(),
                        os_numero=os_numero,
                        cooperado_id=c.id,
                        estabelecimento_id=est.id,
                        valor=valor_f,
                        descricao=descricao
                    )
                    db.session.add(l)
                    c.credito = novo_credito
                    db.session.commit()
                    _update_last_lanc_cache_with_value(l.id)
                    flash('Lançamento realizado com sucesso!', 'success')
            else:
                flash('Cooperado não encontrado!', 'danger')
        else:
            flash('Preencha todos os campos obrigatórios!', 'danger')

    lancamentos = Lancamento.query.filter_by(estabelecimento_id=est.id).order_by(Lancamento.data.desc()).all()

    # Timezone conversion
    try:
        from pytz import timezone, utc
        tz_sp = timezone('America/Sao_Paulo')
        for l in lancamentos:
            hora_brasilia = l.data.replace(tzinfo=utc).astimezone(tz_sp)
            l.data_brasilia = hora_brasilia.strftime('%d/%m/%Y %H:%M')
    except Exception:
        for l in lancamentos:
            l.data_brasilia = l.data.strftime('%d/%m/%Y %H:%M')

    return render_template('painel_estabelecimento.html', est=est, cooperados=cooperados, lancamentos=lancamentos)

# ========= ESTAB: EDITAR / EXCLUIR LANÇAMENTO (10h de janela) =========
@app.route('/estab/lancamento/editar/<int:id>', methods=['POST'])
def estab_editar_lancamento(id):
    if not is_estabelecimento():
        return redirect(url_for('login'))

    l = Lancamento.query.get_or_404(id)

    if l.estabelecimento_id != session.get('user_id'):
        flash('Você não tem permissão para editar este lançamento.', 'danger')
        return redirect(url_for('painel_estabelecimento'))

    if datetime.utcnow() - l.data > timedelta(hours=10):
        flash('Edição permitida somente até 10 horas após a criação.', 'warning')
        return redirect(url_for('painel_estabelecimento'))

    os_numero = request.form.get('os_numero', '').strip()
    valor_str = request.form.get('valor', '').strip().replace(',', '.')
    descricao = request.form.get('descricao', '').strip()

    try:
        novo_valor = float(valor_str)
        if novo_valor <= 0:
            raise ValueError()
    except Exception:
        flash('Valor inválido.', 'danger')
        return redirect(url_for('painel_estabelecimento'))

    if not os_numero:
        flash('O número da OS é obrigatório.', 'danger')
        return redirect(url_for('painel_estabelecimento'))

    cooperado = Cooperado.query.get(l.cooperado_id)
    if not cooperado:
        flash('Cooperado não encontrado.', 'danger')
        return redirect(url_for('painel_estabelecimento'))

    valor_antigo = l.valor
    delta = novo_valor - valor_antigo

    if delta > 0 and cooperado.credito < delta:
        flash('Crédito insuficiente para aumentar o valor deste lançamento.', 'danger')
        return redirect(url_for('painel_estabelecimento'))

    l.os_numero = os_numero
    l.valor = novo_valor
    l.descricao = descricao if descricao else None

    cooperado.credito -= delta
    db.session.commit()
    _invalidate_last_lanc_cache()
    flash('Lançamento editado com sucesso!', 'success')
    return redirect(url_for('painel_estabelecimento'))

@app.route('/estab/lancamento/excluir/<int:id>', methods=['POST'])
def estab_excluir_lancamento(id):
    if not is_estabelecimento():
        return redirect(url_for('login'))

    l = Lancamento.query.get_or_404(id)

    if l.estabelecimento_id != session.get('user_id'):
        flash('Você não tem permissão para excluir este lançamento.', 'danger')
        return redirect(url_for('painel_estabelecimento'))

    if datetime.utcnow() - l.data > timedelta(hours=10):
        flash('Exclusão permitida somente até 10 hora após a criação.', 'warning')
        return redirect(url_for('painel_estabelecimento'))

    cooperado = Cooperado.query.get(l.cooperado_id)
    if not cooperado:
        flash('Cooperado não encontrado.', 'danger')
        return redirect(url_for('painel_estabelecimento'))

    cooperado.credito += l.valor

    db.session.delete(l)
    db.session.commit()
    _invalidate_last_lanc_cache()
    flash('Lançamento excluído e crédito devolvido ao cooperado.', 'success')
    return redirect(url_for('painel_estabelecimento'))

# ========= PAINEL COOPERADO =========
@app.route('/painel_cooperado')
def painel_cooperado():
    if not is_cooperado():
        return redirect(url_for('login'))

    coop = Cooperado.query.get_or_404(session['user_id'])

    # filtros de período (opcionais)
    di_s = request.args.get('data_inicio') or ''
    df_s = request.args.get('data_fim') or ''
    di = parse_date(di_s)
    df = parse_date(df_s)

    q = Lancamento.query.filter(Lancamento.cooperado_id == coop.id)
    if di:
        q = q.filter(Lancamento.data >= di)
    if df:
        q = q.filter(Lancamento.data <= df)
    lancamentos = q.order_by(Lancamento.data.desc()).all()

    # Horário em Brasília
    try:
        from pytz import timezone, utc
        tz_sp = timezone('America/Sao_Paulo')
        for l in lancamentos:
            l.data_brasilia = l.data.replace(tzinfo=utc).astimezone(tz_sp).strftime('%d/%m/%Y %H:%M')
    except Exception:
        for l in lancamentos:
            l.data_brasilia = l.data.strftime('%d/%m/%Y %H:%M')

    total_gasto = sum(float(l.valor or 0) for l in lancamentos)
    total_lanc = len(lancamentos)

    # Tenta usar o template; se não existir no deploy, usa fallback embutido
    try:
        return render_template(
            'painel_cooperado.html',
            coop=coop,
            lancamentos=lancamentos,
            total_gasto=total_gasto,
            total_lanc=total_lanc,
            data_inicio=di_s,
            data_fim=df_s
        )
    except TemplateNotFound:
        # Fallback com layout branco + azul royal (mobile-first)
        return render_template_string("""
<!doctype html>
<html lang="pt-br">
<head>
  <meta charset="utf-8">
  <title>Painel do Cooperado</title>
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700;800&display=swap" rel="stylesheet">
  <style>
    :root{--royal:#0d3ccf;--royal-600:#0b31a9;--royal-50:#e9f0ff;--ink:#111827;--muted:#6b7280;--ok:#13b981;--bg:#f6f8ff;--card:#ffffff;--line:#eef2ff;--chip:#edf2ff}
    *{box-sizing:border-box}html,body{height:100%}
    body{margin:0;background:var(--bg);color:var(--ink);font-family:Inter,system-ui,-apple-system,Segoe UI,Roboto,Arial,sans-serif}
    header{position:sticky;top:0;z-index:20;background:linear-gradient(110deg,var(--royal),var(--royal-600));color:#fff;padding:14px 16px;display:flex;align-items:center;gap:12px;box-shadow:0 6px 16px rgba(0,0,0,.18)}
    .avatar{width:44px;height:44px;border-radius:50%;object-fit:cover;background:#fff2;border:2px solid #ffffff44}
    .title{margin:0;font-weight:800;font-size:18px;line-height:1.1}.subtitle{margin:2px 0 0 0;font-size:12px;opacity:.9}.spacer{flex:1}
    .logout{color:#fff;text-decoration:none;font-weight:700;font-size:13px;border:1px solid #ffffff66;padding:6px 10px;border-radius:10px}.logout:hover{background:#ffffff1a}
    .wrap{max-width:1100px;margin:16px auto;padding:0 12px}
    .grid{display:grid;gap:12px;grid-template-columns:repeat(12,1fr)}
    .card{background:var(--card);border:1px solid; border-color:var(--line);border-radius:16px;padding:14px;box-shadow:0 10px 22px rgba(17,24,39,.04)}
    .card h3{margin:0 0 8px 0;font-size:14px;color:#111827}.metric{font-size:28px;font-weight:800}.metric.ok{color:var(--ok)}.muted{color:var(--muted)}
    .badge{background:var(--chip);color:var(--royal);font-weight:800;padding:4px 10px;border-radius:999px;font-size:12px}
    .filters{display:flex;gap:10px;flex-wrap:wrap;align-items:end;margin:8px 0 12px}
    .filters .field{display:flex;flex-direction:column;gap:6px}.filters label{font-size:12px;color:#374151;font-weight:600}
    .filters input[type=date]{padding:10px 12px;height:42px;border:1px solid var(--line);border-radius:12px;background:#fff}
    .filters button{height:42px;padding:0 16px;border:0;border-radius:12px;font-weight:800;color:#fff;background:linear-gradient(90deg,var(--royal),var(--royal-600));cursor:pointer}
    .filters a.reset{height:42px;display:inline-flex;align-items:center;justify-content:center;padding:0 14px;border-radius:12px;font-weight:700;color:var(--royal-600);background:#e9f0ff;text-decoration:none;border:1px solid var(--line)}
    .table-wrap{overflow:auto;border:1px solid var(--line);border-radius:14px}
    table{width:100%;border-collapse:collapse}th,td{padding:12px 10px;border-bottom:1px solid var(--line);font-size:14px}
    thead th{position:sticky;top:0;background:#f8faff;text-transform:uppercase;letter-spacing:.06em;font-size:12px;color:#4b5563;text-align:left}
    td.right{text-align:right}tr:hover td{background:#fbfdff}.empty{padding:18px;color:var(--muted);text-align:center}
    @media (max-width:900px){.grid .span-4{grid-column:span 12}.grid .span-12{grid-column:span 12}.metric{font-size:24px}}
    @media (min-width:901px){.grid .span-4{grid-column:span 4}.grid .span-12{grid-column:span 12}}
  </style>
</head>
<body>
<header>
  <img class="avatar" src="{{ url_for('foto_cooperado', id=coop.id) }}" alt="foto do cooperado" onerror="this.style.visibility='hidden'">
  <div><h1 class="title">Olá, {{ coop.nome }}</h1><div class="subtitle">Usuário: <b>@{{ coop.username }}</b></div></div>
  <div class="spacer"></div><a class="logout" href="{{ url_for('logout') }}">Sair</a>
</header>
<div class="wrap">
  <div class="grid">
    <div class="card span-4">
      <h3>Crédito disponível</h3>
      <div class="metric ok">R$ {{ '%.2f'|format(coop.credito or 0) }}</div>
      {% if coop.credito_atualizado_em %}
        <div class="muted" style="margin-top:6px">Último ajuste: {{ coop.credito_atualizado_em.strftime('%d/%m/%Y %H:%M') }}</div>
      {% else %}
        <div class="muted" style="margin-top:6px">Sem ajustes manuais registrados</div>
      {% endif %}
    </div>
    <div class="card span-4">
      <h3>Total de lançamentos (período)</h3>
      <div class="metric">{{ total_lanc }}</div>
      <div class="muted" style="margin-top:6px">Itens listados abaixo</div>
    </div>
    <div class="card span-4">
      <h3>Gasto no período</h3>
      <div class="metric">R$ {{ '%.2f'|format(total_gasto or 0) }}</div>
      <div class="muted" style="margin-top:6px">Soma dos lançamentos filtrados</div>
    </div>
    <div class="card span-12">
      <div style="display:flex;align-items:center;justify-content:space-between;gap:12px">
        <h3 style="margin:0">Meus lançamentos</h3>
        <span class="badge">Atualizado agora</span>
      </div>
      <form class="filters" method="get" action="{{ url_for('painel_cooperado') }}">
        <div class="field"><label for="di">Data início</label><input id="di" type="date" name="data_inicio" value="{{ data_inicio }}"></div>
        <div class="field"><label for="df">Data fim</label><input id="df" type="date" name="data_fim" value="{{ data_fim }}"></div>
        <div class="field"><button type="submit">Aplicar filtros</button></div>
        {% if data_inicio or data_fim %}<a class="reset" href="{{ url_for('painel_cooperado') }}">Limpar</a>{% endif %}
      </form>
      <div class="table-wrap">
        <table>
          <thead><tr><th>Data</th><th>Nº OS</th><th>Estabelecimento</th><th class="right">Valor (R$)</th><th>Descrição</th></tr></thead>
          <tbody>
          {% if lancamentos %}
            {% for l in lancamentos %}
              <tr>
                <td>{{ l.data_brasilia or l.data.strftime('%d/%m/%Y %H:%M') }}</td>
                <td>{{ l.os_numero }}</td>
                <td>{{ l.estabelecimento.nome if l.estabelecimento else '-' }}</td>
                <td class="right">{{ '%.2f'|format(l.valor or 0) }}</td>
                <td>{{ l.descricao or '' }}</td>
              </tr>
            {% endfor %}
          {% else %}
            <tr><td colspan="5" class="empty">Nenhum lançamento neste período.</td></tr>
          {% endif %}
          </tbody>
        </table>
      </div>
      <div class="muted" style="margin-top:10px">* Os horários são mostrados em Brasília (GMT-3).</div>
    </div>
  </div>
</div>
</body>
</html>
        """, coop=coop, lancamentos=lancamentos, total_gasto=total_gasto,
           total_lanc=total_lanc, data_inicio=di_s, data_fim=df_s)

# ========= CRIA BANCO + ADMIN MASTER =========
def criar_banco_e_admin():
    with app.app_context():
        db.create_all()
        ensure_schema()
        if not Admin.query.filter_by(username='coopex').first():
            admin = Admin(nome='Administrador Master', username='coopex')
            admin.set_senha('coopex05289')
            db.session.add(admin)
            db.session.commit()
            print('Admin criado: coopex / coopex05289')

# ========= MAIN =========
if __name__ == '__main__':
    criar_banco_e_admin()
    app.run(debug=False, host="0.0.0.0", port=int(os.environ.get("PORT", 5000)))
